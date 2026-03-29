"""
sample_for_coding.py
--------------------
inter-rater reliability용 50건 층화추출 (v2)

층화 설계:
  그룹(3) × 이벤트 전후(2) = 6셀 × 8~9건
  - FM_Direct  : 키워드 그룹 1 (백신 직접 관련)
  - Court      : 키워드 그룹 2 (법원 판결)
  - Chronic    : 키워드 그룹 3 (만성 불신)
  - pre        : 2026-02-07 ~ 2026-02-22
  - post       : 2026-02-23 ~ 2026-03-21

출력:
  data/exports/kappa/sample_50_<ts>.csv            — 분류용 원본
  data/exports/kappa/coding_sample_v2_<ts>.xlsx    — 코더 전달용 폼

Usage:
    python sample_for_coding.py
    python sample_for_coding.py --seed 99   # 재현성 시드 변경
    python sample_for_coding.py --n 50      # 추출 건수 (기본값 50)
"""

import sqlite3
import csv
import random
import argparse
import sys
from pathlib import Path
from datetime import datetime, date

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from pipeline.config import DB_PATH, KAPPA_DIR

# ── 설정 ──────────────────────────────────────────────────────────────────────

NAVER_DB   = DB_PATH
OUTPUT_DIR = KAPPA_DIR

EVENT_DATE = date(2026, 2, 23)   # BAI audit 공개일

# 셀별 목표 건수 (총 50건, 6셀)
CELL_TARGETS = {
    ("FM_Direct", "pre"):  8,
    ("FM_Direct", "post"): 8,
    ("Court",     "pre"):  9,
    ("Court",     "post"): 9,
    ("Chronic",   "pre"):  8,
    ("Chronic",   "post"): 8,
}

# ── 그룹 정의 (키워드 → 그룹 매핑) ───────────────────────────────────────────
# collect_all.py의 KEYWORD_GROUPS와 동일하게 유지

KEYWORD_TO_GROUP = {
    # FM_Direct
    "백신 부작용":    "FM_Direct",
    "코로나 백신":    "FM_Direct",
    "백신 이상반응":  "FM_Direct",
    "mRNA 백신":      "FM_Direct",
    "화이자 부작용":  "FM_Direct",
    "모더나 부작용":  "FM_Direct",
    "백신 피해":      "FM_Direct",
    "백신 사망":      "FM_Direct",
    "백신 후유증":    "FM_Direct",
    "아나필락시스":   "FM_Direct",
    # Court
    "백신 소송":      "Court",
    "백신 보상":      "Court",
    "예방접종 피해":  "Court",
    "심근염 백신":    "Court",
    "백신 국가배상":  "Court",
    "백신 판결":      "Court",
    "백신 승소":      "Court",
    # Chronic
    "질병관리청 백신": "Chronic",
    "백신 불신":       "Chronic",
    "백신 거부":       "Chronic",
    # Political → FM_Direct (reclassified Week 8)
    "정은경 백신":     "FM_Direct",
}


def get_post_group(row: dict) -> str:
    """keyword_group 컬럼 우선 사용, 없으면 keyword 매핑."""
    if row.get("keyword_group"):
        return row["keyword_group"]
    return KEYWORD_TO_GROUP.get(row.get("keyword", ""), "Unknown")


def get_post_period(pub_date: str) -> str:
    """pub_date(YYYY-MM-DD) → 'pre' or 'post'."""
    try:
        d = date.fromisoformat(pub_date[:10])
        return "pre" if d < EVENT_DATE else "post"
    except Exception:
        return "unknown"


def load_posts(db_path: Path) -> list[dict]:
    """naver_posts.db에서 전체 포스트 로드."""
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.execute("""
        SELECT post_id, keyword, keyword_group, channel, content,
               published_at AS pub_date
        FROM posts
        WHERE published_at IS NOT NULL
          AND published_at != ''
          AND is_relevant = 1
    """)
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows


def deduplicate_by_content(posts: list[dict]) -> list[dict]:
    """content 기준 텍스트 중복 제거."""
    seen = set()
    result = []
    for p in posts:
        key = (p.get("content") or "").strip()
        if key and key not in seen:
            seen.add(key)
            result.append(p)
    return result


def stratified_sample(posts: list[dict], targets: dict, seed: int) -> list[dict]:
    """6셀 층화추출. 각 셀에서 무작위 추출."""
    rng = random.Random(seed)

    # 셀별로 포스트 분류
    cells: dict[tuple, list] = {k: [] for k in targets}
    skipped = 0
    for p in posts:
        group  = get_post_group(p)
        period = get_post_period(p["pub_date"])
        key    = (group, period)
        if key in cells:
            cells[key].append(p)
        else:
            skipped += 1

    print(f"\n[셀별 풀 크기]")
    for key, pool in cells.items():
        print(f"  {key[0]:12s} × {key[1]:4s} : {len(pool):4d}건 풀 → {targets[key]}건 추출")

    if skipped:
        print(f"  (그룹 미매핑 포스트 {skipped}건 제외)")

    # 추출
    sampled = []
    for key, target in targets.items():
        pool = cells[key]
        if len(pool) < target:
            print(f"  ⚠️  {key} 풀 부족({len(pool)}건) — 전량 사용")
            chosen = pool
        else:
            chosen = rng.sample(pool, target)
        for p in chosen:
            p["group"]  = key[0]
            p["period"] = key[1]
        sampled.extend(chosen)

    # 순서 셔플 (코더가 그룹 패턴 파악 못하도록)
    rng.shuffle(sampled)
    for i, p in enumerate(sampled, 1):
        p["sample_id"] = i

    return sampled


def save_csv(samples: list[dict], out_path: Path) -> None:
    """분석용 CSV 저장."""
    fields = ["sample_id", "group", "period", "channel",
              "keyword", "pub_date", "content", "post_id"]
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(samples)
    print(f"\n✅ CSV 저장: {out_path}")


def save_coding_form(samples: list[dict], out_path: Path) -> None:
    """코더 전달용 xlsx 코딩 폼 생성."""
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── 시트 1: 코딩 폼 ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "코딩폼"

    # 색상 정의
    HDR_FILL  = PatternFill("solid", start_color="1F4E79")   # 진청
    META_FILL = PatternFill("solid", start_color="D6E4F0")   # 연청
    CODE_FILL = PatternFill("solid", start_color="FFF2CC")   # 연노랑
    DONE_FILL = PatternFill("solid", start_color="E2EFDA")   # 연초록
    THIN = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    # 헤더
    headers = [
        "sample_id", "group", "period", "channel", "pub_date",
        "제목/내용",
        "C1\n(Confidence)", "C2\n(Complacency)", "C4\n(Calculation)",
        "C6\n(Compliance)", "C7\n(Conspiracy)",
        "코딩 메모"
    ]
    ws.append(headers)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = THIN
    ws.row_dimensions[1].height = 36

    # 데이터 행
    for row_idx, p in enumerate(samples, 2):
        # 텍스트: 내용 앞 200자
        content_preview = (p.get("content") or "").strip()[:200]
        text_cell       = content_preview + ("…" if len((p.get("content") or "")) > 200 else "")

        row_data = [
            p["sample_id"],
            p["group"],
            p["period"],
            p["channel"],
            p["pub_date"][:10] if p["pub_date"] else "",
            text_cell,
            "", "", "", "", "",   # C1~C7 입력란
            ""                    # 메모
        ]
        ws.append(row_data)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = THIN
            cell.font   = Font(name="Arial", size=9)
            if col <= 5:
                cell.fill      = META_FILL
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col == 6:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif col <= 11:
                cell.fill      = CODE_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.row_dimensions[row_idx].height = 72

    # 열 너비
    col_widths = [8, 12, 7, 14, 11, 60, 8, 8, 8, 8, 8, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 틀 고정
    ws.freeze_panes = "G2"

    # ── 시트 2: 코딩 가이드 ──────────────────────────────────────────────────
    guide = wb.create_sheet("코딩가이드")
    guide_rows = [
        ["Infovail-IQ PoC1 — 7C 코딩 가이드", "", "", ""],
        ["버전: v1.0  |  기준 프롬프트: v1.1.0  |  작성: Suah Cheon, MD", "", "", ""],
        [""],
        ["[입력 방법]", "", "", ""],
        ["• C1~C7 열에 1(해당) 또는 0(비해당)을 입력하세요.", "", "", ""],
        ["• 복합 레이블 가능 (예: C1=1, C6=1, C7=1 동시 가능)", "", "", ""],
        ["• 코딩 메모: 판단이 어려운 경우 근거를 간략히 기재", "", "", ""],
        ["• 비어 있으면 미코딩으로 처리됩니다.", "", "", ""],
        [""],
        ["[차원 정의]", "", "", ""],
        ["차원", "이름", "핵심 정의", "한국어 예시"],
        ["C1", "Confidence\n(신뢰 결여)",
         "백신 안전성·효능 또는 보건당국/제약사/규제기관에 대한 불신",
         "\"이물질이 든 백신을 1420만 명한테 맞혔다. 식약처는 뭐했냐\""],
        ["C2", "Complacency\n(안주)",
         "백신 예방 질병의 위험을 낮게 인식해 접종이 불필요하다고 판단",
         "\"이제 코로나 끝났는데 아직도 맞으라고? 그냥 감기야\""],
        ["C4", "Calculation\n(계산)",
         "접종 여부 결정 전 정보를 적극적으로 탐색하거나 손익을 따지는 행동",
         "\"부작용 데이터가 좀 더 나오면 그때 판단하려고요\""],
        ["C6", "Compliance\n(순응 저항)",
         "강제 접종 또는 접종 압박에 대한 저항",
         "\"방역패스는 미접종자 차별이다. 헌법 위반이야\""],
        ["C7", "Conspiracy\n(음모론)",
         "백신 정책 뒤에 악의적 의도가 있다는 믿음 — 단순 불신 이상",
         "\"제약사가 정부에 로비해서 허가받은 거지. 다 짜고 치는 고스톱\""],
        [""],
        ["[핵심 구별 기준]", "", "", ""],
        ["C1 vs C7",
         "C1 = 무능/부주의 가능성 포함한 불신  |  C7 = 의도적 기만을 명시적으로 주장",
         "", ""],
        ["C1 vs C4",
         "C1 = 불신 표현  |  C4 = '더 알아보고 결정하겠다'는 의도 명시",
         "", ""],
        ["C1 vs C6",
         "C1 = 백신/시스템 불신  |  C6 = 강제 접종에 대한 저항 (백신 자체 문제 아닐 수 있음)",
         "", ""],
        [""],
        ["[주의사항]", "", "", ""],
        ["• 판단 불가능한 항목은 0으로 남기고 메모란에 '판단불가' 기재", "", "", ""],
        ["• 코딩 완료 후 파일명에 코더 이니셜 추가 (예: coding_form_XXX_CoderB.xlsx)", "", "", ""],
        ["• 다른 코더의 결과를 보기 전에 독립적으로 코딩 완료할 것", "", "", ""],
    ]
    for r in guide_rows:
        guide.append(r)

    # 가이드 시트 헤더 스타일
    for col in range(1, 5):
        cell = guide.cell(row=1, column=col)
        cell.font = Font(bold=True, size=13, name="Arial", color="1F4E79")
    guide.cell(row=11, column=1).font = Font(bold=True, name="Arial")
    for row in [11]:
        for col in range(1, 5):
            guide.cell(row=row, column=col).fill = PatternFill("solid", start_color="D6E4F0")

    guide.column_dimensions["A"].width = 12
    guide.column_dimensions["B"].width = 22
    guide.column_dimensions["C"].width = 55
    guide.column_dimensions["D"].width = 45

    # ── 시트 3: 추출 메타데이터 ──────────────────────────────────────────────
    meta = wb.create_sheet("추출메타")
    meta.append(["항목", "값"])
    meta.append(["추출일시", datetime.now().strftime("%Y-%m-%d %H:%M")])
    meta.append(["총 추출건수", len(samples)])
    meta.append(["층화 기준", "키워드그룹(3) × 이벤트전후(2) = 6셀 × 8~9건"])
    meta.append(["이벤트 기준일", str(EVENT_DATE)])
    meta.append(["랜덤 시드", "기록됨 (재현 가능)"])
    meta.append([""])
    meta.append(["셀", "목표", "실제"])
    for key, target in CELL_TARGETS.items():
        actual = sum(1 for s in samples if s["group"] == key[0] and s["period"] == key[1])
        meta.append([f"{key[0]} × {key[1]}", target, actual])

    wb.save(out_path)
    print(f"✅ 코딩 폼 저장: {out_path}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--seed", type=int, default=42,
                        help="랜덤 시드 (기본값: 42)")
    parser.add_argument("--n", type=int, default=50,
                        help="추출 건수 (기본값: 50)")
    args = parser.parse_args()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    print(f"[Infovail-IQ] 층화추출 시작 - seed={args.seed}, n={args.n}")
    print(f"DB: {NAVER_DB}")

    posts   = load_posts(NAVER_DB)
    print(f"전체 포스트: {len(posts)}건")

    posts = deduplicate_by_content(posts)
    print(f"중복 제거 후: {len(posts)}건")

    samples = stratified_sample(posts, CELL_TARGETS, seed=args.seed)
    print(f"\n추출 완료: {len(samples)}건")

    csv_path  = OUTPUT_DIR / f"sample_50_{ts}.csv"
    xlsx_path = OUTPUT_DIR / f"coding_sample_v2_{ts}.xlsx"

    save_csv(samples, csv_path)
    save_coding_form(samples, xlsx_path)

    print(f"\n[다음 단계]")
    print(f"  1. {xlsx_path.name} → Google Sheets 업로드 → 코더 B 공유")
    print(f"  2. Suah(코더 A)도 동일 파일 독립 코딩")
    print(f"  3. 두 코딩 완료 후 kappa 계산 스크립트 실행")


if __name__ == "__main__":
    main()
